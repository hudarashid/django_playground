from blog.resources import BookResource
from django.contrib import admin
from import_export.admin import ExportActionMixin
from .models import Blog, Category

from etc.admin import admins, CustomModelPage
from django.db import models
from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook


class CategoryAdmin(admin.ModelAdmin):
    model = Category

    list_display = ("name",)


class BlogAdmin(ExportActionMixin, admin.ModelAdmin):

    model = Blog

    resource_classes = [BookResource]

    list_display = (
        "title",
        "body",
        "published_date",
        "created_date",
        "edited_date",
        "is_published",
    )

    field = (
        "title",
        "body",
        "blog_category",
        "published_date",
        "is_published",
    )

class BlogExportModelAdmin(admins.CustomPageModelAdmin):
    # declare the fields to be exported
    fields = (
        "blog_title",
        "body",
        "created_date",
        "published_date",
        "edited_date",
        "is_published",
        
    )

class BlogExportPage(CustomModelPage):

    title = 'Blog Export Option'

    # Fields you want to proccess data from.
    blog_title = models.BooleanField(default=False)
    body = models.BooleanField(default=False)
    published_date = models.BooleanField(default=False)
    created_date = models.BooleanField(default=False)
    edited_date = models.BooleanField(default=False)
    is_published = models.BooleanField(default=False)

    admin_cls = BlogExportModelAdmin  

    def save(self):
        # Get the selected fields from the Blog model
        selected_fields = [field.name for field in Blog._meta.get_fields() if getattr(self, field.name, False)]
     
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="blog_export-{datetime.now().strftime("%Y-%m-%d")}.xlsx"'

        workbook = Workbook()
        worksheet = workbook.active

        # Write the header row
        for col_num, column_title in enumerate(selected_fields, 1):
            worksheet.cell(row=1, column=col_num, value=column_title)

        # Write the data rows
        objects = Blog.objects.values(*selected_fields)
        for row_num, obj in enumerate(objects, 2):
            for col_num, field in enumerate(selected_fields, 1):
                worksheet.cell(row=row_num, column=col_num, value=str(obj[field]))

        workbook.save(response)

        self.bound_response = response

        super().save()

# Register my page within Django admin.
BlogExportPage.register()



admin.site.register(Blog, BlogAdmin)
admin.site.register(Category, CategoryAdmin)